from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook, load_workbook

from src.domain.schemas import ExtractRow, WriteSummary

INVALID_FILENAME_CHARS_PATTERN = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
MAX_EXPORT_FILENAME_LENGTH = 120
DEFAULT_OUTPUT_FILENAME = "example.xlsx"


class ExcelWriter:
    def __init__(self, project_root: str | Path | None = None) -> None:
        self.project_root = Path(project_root or Path.cwd())
        self.last_output_resolution: dict[str, str] | None = None

    def append_rows(self, output_path: str, rows: list[list[str]]) -> WriteSummary:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        if path.exists():
            workbook = load_workbook(path)
            worksheet = workbook.worksheets[0]
        else:
            workbook = Workbook()
            worksheet = workbook.worksheets[0]

        for row in rows:
            worksheet.append(row)

        workbook.save(path)
        return WriteSummary(
            total_rows=len(rows),
            selected_rows=len(rows),
            written_rows=len(rows),
            skipped_rows=0,
            failed_rows=0,
            output_path=str(path),
        )

    def write_result_rows(self, result_rows: list[ExtractRow], output_path: str) -> WriteSummary:
        review_rows = [
            {
                "row_id": row.row_id,
                "task_id": row.task_id,
                "values": list(row.values),
                "typed_values": list(row.typed_values) if row.typed_values is not None else None,
                "action": row.action,
                "is_error_row": row.is_error_row,
            }
            for row in result_rows
        ]
        return self.write_review_rows(review_rows, output_path)

    def write_review_rows(self, rows: list[Mapping[str, object] | Any], output_path: str) -> WriteSummary:
        effective_output = output_path or str(self.project_root / DEFAULT_OUTPUT_FILENAME)
        resolved_output_path, resolution = self._resolve_output_path(effective_output)
        self.last_output_resolution = resolution
        selected = [row for row in rows if _row_action(row) == "write" and not _row_is_error(row)]
        payload = [_row_typed_values(row) if _row_typed_values(row) is not None else _row_values(row) for row in selected]

        write_summary = self.append_rows(str(resolved_output_path), payload)
        return WriteSummary(
            total_rows=len(rows),
            selected_rows=len(selected),
            written_rows=write_summary.written_rows,
            skipped_rows=len(rows) - len(selected),
            failed_rows=write_summary.failed_rows,
            output_path=write_summary.output_path,
        )

    def _resolve_output_path(self, output_path: str) -> tuple[Path, dict[str, str]]:
        raw_path = Path(output_path).expanduser()
        parent = raw_path.parent if str(raw_path.parent).strip() not in {"", "."} else self.project_root
        original_filename = raw_path.name or DEFAULT_OUTPUT_FILENAME
        normalized_filename = self._normalize_filename(original_filename)
        resolved_path = parent / normalized_filename
        return resolved_path, {
            "original_output_path": str(raw_path),
            "original_filename": original_filename,
            "normalized_filename": normalized_filename,
            "resolved_output_path": str(resolved_path),
            "resolved_filename": resolved_path.name,
        }

    @staticmethod
    def _normalize_filename(filename: str) -> str:
        raw_name = str(filename or "").strip()
        if not raw_name:
            raw_name = DEFAULT_OUTPUT_FILENAME
        raw_path = Path(raw_name)
        suffix = raw_path.suffix or ".xlsx"
        stem = raw_path.stem if raw_path.suffix else raw_name
        cleaned_stem = INVALID_FILENAME_CHARS_PATTERN.sub("", stem).strip().rstrip(".")
        if not cleaned_stem:
            cleaned_stem = "export"
        max_stem_len = max(1, MAX_EXPORT_FILENAME_LENGTH - len(suffix))
        if len(cleaned_stem) > max_stem_len:
            cleaned_stem = cleaned_stem[:max_stem_len].rstrip(" ._")
        if not cleaned_stem:
            cleaned_stem = "export"
        return f"{cleaned_stem}{suffix}"

    @staticmethod
    def _dedupe_output_path(candidate_path: Path) -> Path:
        """Retained as rollback anchor for unify-toolbar-and-excel-path; not invoked."""
        if not candidate_path.exists():
            return candidate_path
        stem = candidate_path.stem
        suffix = candidate_path.suffix
        index = 2
        while True:
            deduped_path = candidate_path.with_name(f"{stem}_{index}{suffix}")
            if not deduped_path.exists():
                return deduped_path
            index += 1


def _row_action(row: Mapping[str, object] | Any) -> str:
    if isinstance(row, Mapping):
        return str(row.get("action") or "write")
    return str(getattr(row, "action", "write") or "write")


def _row_is_error(row: Mapping[str, object] | Any) -> bool:
    if isinstance(row, Mapping):
        return bool(row.get("is_error_row", False))
    return bool(getattr(row, "is_error_row", False))


def _row_values(row: Mapping[str, object] | Any) -> list[object]:
    if isinstance(row, Mapping):
        return list(row.get("values") or [])
    return list(getattr(row, "values", []) or [])


def _row_typed_values(row: Mapping[str, object] | Any) -> list[object] | None:
    if isinstance(row, Mapping):
        raw_value = row.get("typed_values")
    else:
        raw_value = getattr(row, "typed_values", None)
    if raw_value is None:
        return None
    return list(raw_value)
