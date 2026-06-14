from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from src.domain.schemas import ExtractRow
from src.io.excel_writer import ExcelWriter
from src.ui.task_review_projection import TaskReviewExportRow


def test_append_rows_creates_file_and_appends_to_first_sheet(tmp_path: Path) -> None:
    path = tmp_path / "out.xlsx"
    writer = ExcelWriter(project_root=tmp_path)

    summary = writer.append_rows(str(path), [["A", "B"], ["C", "D"]])

    assert summary.written_rows == 2
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    assert ws.max_row == 2
    assert ws.cell(1, 1).value == "A"
    assert ws.cell(2, 2).value == "D"


def test_append_rows_appends_existing_sheet_tail(tmp_path: Path) -> None:
    path = tmp_path / "out.xlsx"
    writer = ExcelWriter(project_root=tmp_path)
    writer.append_rows(str(path), [["A", "B"]])

    summary = writer.append_rows(str(path), [["C", "D"]])

    assert summary.written_rows == 1
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "C"


def test_write_result_rows_filters_skip_and_uses_default_output_path(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    rows = [
        ExtractRow(task_id="1", values=["A", "B"], action="write"),
        ExtractRow(task_id="1", values=["X", "Y"], action="skip"),
    ]

    summary = writer.write_result_rows(rows, output_path="")

    assert summary.output_path == str(tmp_path / "example.xlsx")
    assert summary.total_rows == 2
    assert summary.selected_rows == 1
    assert summary.written_rows == 1
    assert summary.skipped_rows == 1


def test_write_result_rows_auto_normalizes_chinese_filename_with_invalid_chars(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    rows = [ExtractRow(task_id="1", values=["A", "B"], action="write")]

    summary = writer.write_result_rows(rows, output_path=str(tmp_path / "中文:报告*?.xlsx"))

    resolved_path = Path(summary.output_path)
    assert resolved_path.exists()
    assert resolved_path.name == "中文报告.xlsx"


def test_normalize_filename_strips_trailing_dots_and_spaces() -> None:
    assert ExcelWriter._normalize_filename("report. .xlsx") == "report.xlsx"


def test_normalize_filename_uses_default_for_blank_name() -> None:
    assert ExcelWriter._normalize_filename("   ") == "example.xlsx"


def test_normalize_filename_uses_export_when_invalid_chars_empty_stem() -> None:
    assert ExcelWriter._normalize_filename("<>:\"/\\|?*.xlsx") == "export.xlsx"


def test_write_result_rows_appends_in_place_to_existing_target_keeping_single_sheet(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    target_path = tmp_path / "exist.xlsx"
    writer.append_rows(str(target_path), [["old_1a", "old_1b"], ["old_2a", "old_2b"], ["old_3a", "old_3b"]])

    rows = [
        ExtractRow(task_id="t", values=["new_1a", "new_1b"], action="write"),
        ExtractRow(task_id="t", values=["new_2a", "new_2b"], action="write"),
    ]
    summary = writer.write_result_rows(rows, output_path=str(target_path))

    assert summary.output_path == str(target_path)
    wb = load_workbook(target_path)
    assert len(wb.worksheets) == 1
    ws = wb.worksheets[0]
    assert ws.max_row == 5
    assert ws.cell(1, 1).value == "old_1a"
    assert ws.cell(3, 2).value == "old_3b"
    assert ws.cell(4, 1).value == "new_1a"
    assert ws.cell(5, 2).value == "new_2b"


def test_write_result_rows_called_twice_appends_in_place_without_creating_dedup_suffix(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    target_path = tmp_path / "twice.xlsx"
    rows = [ExtractRow(task_id="t", values=["a", "b"], action="write")]

    writer.write_result_rows(rows, output_path=str(target_path))
    writer.write_result_rows(rows, output_path=str(target_path))

    assert (target_path).exists()
    assert not (tmp_path / "twice_2.xlsx").exists()
    wb = load_workbook(target_path)
    assert len(wb.worksheets) == 1
    assert wb.worksheets[0].max_row == 2


def test_write_result_rows_default_path_called_twice_appends_to_singular_example_xlsx(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    rows = [ExtractRow(task_id="t", values=["a", "b"], action="write")]

    writer.write_result_rows(rows, output_path="")
    writer.write_result_rows(rows, output_path="")

    assert (tmp_path / "example.xlsx").exists()
    assert not (tmp_path / "examples.xlsx").exists()
    assert not (tmp_path / "example_2.xlsx").exists()
    wb = load_workbook(tmp_path / "example.xlsx")
    assert wb.worksheets[0].max_row == 2


def test_write_result_rows_succeeds_when_target_file_exists_with_empty_first_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook

    target_path = tmp_path / "empty_sheet.xlsx"
    Workbook().save(target_path)

    writer = ExcelWriter(project_root=tmp_path)
    rows = [ExtractRow(task_id="t", values=["a", "b"], action="write")]

    summary = writer.write_result_rows(rows, output_path=str(target_path))

    assert summary.output_path == str(target_path)
    wb = load_workbook(target_path)
    assert wb.worksheets[0].cell(1, 1).value == "a"


def test_write_result_rows_prefers_typed_values_for_excel_types(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    rows = [
        ExtractRow(
            task_id="1",
            values=["￥1234.50", "2023年07月03日"],
            typed_values=[1234.5, date(2023, 7, 3)],
            action="write",
        )
    ]

    summary = writer.write_result_rows(rows, output_path=str(tmp_path / "typed.xlsx"))

    wb = load_workbook(summary.output_path)
    ws = wb.worksheets[0]
    assert ws.cell(1, 1).value == 1234.5
    assert ws.cell(1, 1).data_type == "n"
    assert ws.cell(1, 2).value.date().isoformat() == "2023-07-03"


def test_write_review_rows_filters_skip_and_error_and_prefers_typed_values(tmp_path: Path) -> None:
    writer = ExcelWriter(project_root=tmp_path)
    rows = [
        TaskReviewExportRow(
            row_id="row-1",
            task_id="task-1",
            values=["￥1234.50", "2023年07月03日"],
            typed_values=[1234.5, date(2023, 7, 3)],
            action="write",
            is_error_row=False,
        ),
        TaskReviewExportRow(
            row_id="row-2",
            task_id="task-1",
            values=["skip", "me"],
            typed_values=None,
            action="skip",
            is_error_row=False,
        ),
        TaskReviewExportRow(
            row_id="row-3",
            task_id="task-1",
            values=["error", "row"],
            typed_values=["should", "not-write"],
            action="write",
            is_error_row=True,
        ),
    ]

    summary = writer.write_review_rows(rows, output_path=str(tmp_path / "review.xlsx"))

    assert summary.total_rows == 3
    assert summary.selected_rows == 1
    assert summary.skipped_rows == 2
    wb = load_workbook(summary.output_path)
    ws = wb.worksheets[0]
    assert ws.max_row == 1
    assert ws.cell(1, 1).value == 1234.5
    assert ws.cell(1, 2).value.date().isoformat() == "2023-07-03"
